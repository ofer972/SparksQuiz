"""
Import Kahoot-style quizzes from Excel files into SparksQuiz DB.

- Reads all .xlsx files from a directory, one by one (no external file list).
- Each file = one quiz. Each file contains multiple sheets named "1 Quiz", "2 Quiz", "3 Quiz", ...
  Each such sheet = one question, with an "Answer Combinations" section listing answer options
  and which one is correct (checkmark vs X).
- If a quiz with the same title already exists, prompts: Replace? [y/N].

Run from Quiz-Backend with DATABASE_URL in env (e.g. .env):
  python scripts/import_kahoot_quizzes.py [--dir "KahootQuiz"]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

# Run from Quiz-Backend so parent is on path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from sqlalchemy import text

from database_connection import get_connection_string, get_db_engine
from database_quiz import create_quiz, delete_quiz


def get_quiz_id_by_title(conn, title: str) -> int | None:
    row = conn.execute(
        text("SELECT id FROM quizzes WHERE title = :title"),
        {"title": title},
    ).fetchone()
    return row.id if row else None


def _sheet_number(name: str) -> int | None:
    """Extract question number from sheet name: '1 Quiz', '2 Quiz', '1', '2', etc. Return None if not a question sheet."""
    name = (name or "").strip()
    # "1 Quiz", "2 Quiz", "3 Quiz" or just "1", "2", "3"
    if name.isdigit():
        return int(name)
    parts = name.split()
    if len(parts) >= 1 and parts[0].isdigit():
        return int(parts[0])
    if len(parts) >= 2 and parts[1].lower() == "quiz" and parts[0].isdigit():
        return int(parts[0])
    return None


# Kahoot report often has shape icons (triangle, diamond, circle, square) in cells; we want text only
_KAHOOT_SHAPE_CHARS = (
    "\u25B2\u25B3\u25BC\u25BD"  # triangles ▲ △ ▼ ▽
    "\u25C6\u25C7\u25C4\u25C5"  # diamonds ◆ ◇ ◄ ►
    "\u25CF\u25CB\u25E6"        # circles ● ○ ◦
    "\u25A0\u25A1\u25AA\u25AB"  # squares ■ □ ▪ ▫
    "\u2B9E\u2B9F\u2B9D\u2B9C"  # more triangles
)


def _strip_shape_prefix(s: str) -> str:
    """Remove leading Kahoot shape/icon characters; return the text part only."""
    if not s or not isinstance(s, str):
        return (s or "").strip()
    s = s.strip()
    while s and s[0] in _KAHOOT_SHAPE_CHARS:
        s = s[1:].strip()
    while s and s[-1] in _KAHOOT_SHAPE_CHARS:
        s = s[:-1].strip()
    return s.strip()


def _is_answer_text(s: str) -> bool:
    """True if s looks like actual answer text (has letters), not an icon-only cell."""
    if not s or not str(s).strip():
        return False
    t = _strip_shape_prefix(str(s).strip())
    if not t:
        return False
    # Reject if it's only a single character (icon) or has no letters
    if len(t) <= 1:
        return False
    if not any(c.isalpha() for c in t):
        return False
    return True


def _is_correct_cell(value) -> bool:
    """True if the cell indicates the correct answer (checkmark, 'Correct', etc.)."""
    if value is True or value == 1:
        return True
    s = (str(value).strip().lower() if value is not None else "") or ""
    if s in ("✓", "✔", "correct", "yes", "true", "1"):
        return True
    if s in ("×", "✗", "x", "incorrect", "no", "false", "0"):
        return False
    # Checkmark symbols (unicode)
    if value and ord(str(value)[:1]) in (0x2713, 0x2714, 0x2705):
        return True
    # If it looks like a checkmark and not an X, treat as correct
    if s and "correct" in s and "in" not in s:
        return True
    return False


def _parse_question_sheet(ws, sheet_title: str, default_time: int) -> dict | None:
    """
    Parse one worksheet (one question) in Kahoot report format.
    - Question text: same row as "1 Quiz" / "2 Quiz" (sheet title), in the cell to the right.
    - Answers: from "Answer options" row (texts) and "Is answer correct?" row (✓/X).
    Returns a question dict or None.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None

    question_text = None
    answer_options_row = None  # row index where answer option texts are
    is_correct_row = None  # row index where ✓/X are

    for r_idx, row in enumerate(rows[:50]):
        row = list(row) if row is not None else []
        for c_idx, cell in enumerate(row):
            val = str(cell).strip() if cell is not None else ""
            vlo = val.lower()

            # Question: row containing the sheet label (e.g. "2 Quiz"); question text is in same row, next non-empty cell(s)
            if sheet_title and val.strip().lower() == sheet_title.strip().lower():
                for j in range(c_idx + 1, len(row)):
                    other = row[j] if j < len(row) else None
                    if other is not None and str(other).strip():
                        t = str(other).strip()
                        if len(t) > 2 and "answer" not in t.lower() and "correct" not in t.lower():
                            question_text = t[:2000]
                            break
                if question_text:
                    break

            # "Answer options" row: the option texts are in this row (or the next)
            if "answer option" in vlo and "is answer correct" not in vlo:
                # Option texts often in same row (after header) or next row
                answer_options_row = r_idx
            # "Is answer correct?" row: ✓/X in this row (or next)
            if "is answer correct" in vlo:
                is_correct_row = r_idx

    # If we didn't find "Answer options" / "Is answer correct?", try legacy "Answer Combinations" layout
    legacy_combo_row = None
    if answer_options_row is None or is_correct_row is None:
        for r_idx, row in enumerate(rows[:30]):
            row = list(row) if row is not None else []
            for cell in row:
                val = str(cell).strip() if cell is not None else ""
                if "answer combination" in val.lower():
                    legacy_combo_row = r_idx
                    break

    answers = []
    if answer_options_row is not None and is_correct_row is not None:
        # Report format: "Answer options" and "Is answer correct?" are row headers; data is often in the NEXT row
        texts = []
        correct_flags = []
        start_col = 0
        start_c = 0

        # Try data in same row as "Answer options" first
        opts_row = list(rows[answer_options_row]) if answer_options_row < len(rows) else []
        if opts_row and str(opts_row[0]).strip().lower().startswith("answer"):
            start_col = 1
        for c in range(start_col, len(opts_row)):
            v = opts_row[c]
            if v is not None and str(v).strip():
                t = str(v).strip()
                if "answer option" not in t.lower() and "is answer correct" not in t.lower():
                    texts.append(t[:500])

        # If no option texts in same row, use next row (report layout)
        if len(texts) < 2 and answer_options_row + 1 < len(rows):
            opts_row = list(rows[answer_options_row + 1])
            texts = [str(v).strip()[:500] for v in opts_row if v is not None and str(v).strip() and "answer" not in str(v).lower()]

        correct_row = list(rows[is_correct_row]) if is_correct_row < len(rows) else []
        if correct_row and str(correct_row[0]).strip().lower().startswith("is answer"):
            start_c = 1
        for c in range(start_c, min(start_c + len(texts), len(correct_row))):
            correct_flags.append(_is_correct_cell(correct_row[c] if c < len(correct_row) else None))

        if len(correct_flags) < len(texts) and is_correct_row + 1 < len(rows):
            correct_row = list(rows[is_correct_row + 1])
            correct_flags = [_is_correct_cell(correct_row[c] if c < len(correct_row) else None) for c in range(len(texts))]

        # Keep only cells that are real text (not icons); strip leading shape characters
        for i, text in enumerate(texts):
            if not text:
                continue
            cleaned = _strip_shape_prefix(text)
            if not _is_answer_text(cleaned):
                continue
            is_correct = correct_flags[i] if i < len(correct_flags) else False
            answers.append({"answer_text": cleaned[:500], "is_correct": is_correct})

    # Legacy "Answer Combinations" format: each data row has indicator (B) + answer text (C)
    if len(answers) < 2 and legacy_combo_row is not None:
        data_start = legacy_combo_row + 2
        for r_idx in range(data_start, min(data_start + 10, len(rows))):
            row = list(rows[r_idx]) if r_idx < len(rows) else []
            while len(row) < 4:
                row.append(None)
            text_val = row[2] if len(row) > 2 else None
            if text_val is None or not str(text_val).strip():
                continue
            raw = str(text_val).strip()
            if raw.lower() in ("player details", "selected answers", "number of answers"):
                break
            cleaned = _strip_shape_prefix(raw)
            if not _is_answer_text(cleaned):
                continue
            answers.append({
                "answer_text": cleaned[:500],
                "is_correct": _is_correct_cell(row[1] if len(row) > 1 else None),
            })

    if len(answers) < 2:
        return None

    num_correct = sum(1 for a in answers if a["is_correct"])
    question_type = "multi" if num_correct > 1 else "single"
    if not question_text or not question_text.strip():
        question_text = "Question"

    return {
        "question_text": question_text,
        "question_type": question_type,
        "time_limit": default_time,
        "answers": answers,
    }


def parse_excel_to_quiz(file_path: Path, default_time: int = 30) -> tuple[str, list[dict]] | None:
    """
    Read Excel file: each numbered sheet ('1 Quiz', '2 Quiz', ...) = one question.
    Return (quiz_title, questions) or None on error.
    """
    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        return None

    # Load with data_only=True to get values; read_only=False so we can access sheets by name
    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
    # Collect (number, sheet_title) for sheets named "1", "2", "1 Quiz", "2 Quiz", etc.
    numbered = []
    for name in wb.sheetnames:
        n = _sheet_number(name)
        if n is not None:
            numbered.append((n, name))
    numbered.sort(key=lambda x: x[0])

    questions = []
    for _num, sheet_title in numbered:
        ws = wb[sheet_title]
        q = _parse_question_sheet(ws, sheet_title, default_time)
        if q:
            questions.append(q)
    wb.close()

    quiz_title = file_path.stem
    if not questions:
        return None
    return (quiz_title, questions)


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Kahoot-style Excel quizzes into SparksQuiz DB.")
    parser.add_argument(
        "--dir",
        type=str,
        default=None,
        help="Directory containing .xlsx files (default: 'Kahoot Quiz' next to project root)",
    )
    args = parser.parse_args()

    if args.dir:
        dir_path = Path(args.dir).resolve()
    else:
        # Default: project root / KahootQuiz (e.g. C:\SparksQuiz\KahootQuiz)
        project_root = Path(__file__).resolve().parent.parent.parent
        dir_path = project_root / "KahootQuiz"

    if not dir_path.is_dir():
        print(f"Directory not found: {dir_path}", file=sys.stderr)
        sys.exit(1)

    engine = get_db_engine()
    count_created = 0
    count_skipped = 0
    count_replaced = 0

    # Iterate files one by one (no pre-built list required; we just glob and process each)
    for file_path in sorted(dir_path.glob("*.xlsx")):
        result = parse_excel_to_quiz(file_path)
        if not result:
            print(f"Skipped (parse error or empty): {file_path.name}")
            count_skipped += 1
            continue

        title, questions = result

        for i, q in enumerate(questions, 1):
            print(f"  Q{i}: {q['question_text'][:80]}{'...' if len(q['question_text']) > 80 else ''} ({len(q['answers'])} answers, {q['question_type']})")

        with engine.begin() as conn:
            existing_id = get_quiz_id_by_title(conn, title)
            if existing_id is not None:
                answer = input(f"Quiz '{title}' already exists. Replace? [y/N]: ").strip().lower()
                if answer in ("y", "yes"):
                    delete_quiz(conn, existing_id)
                    create_quiz(conn, title, "Imported from Excel", questions)
                    count_replaced += 1
                    print(f"Replaced: {title} ({len(questions)} questions)")
                else:
                    print(f"Skipped (keep existing): {title}")
                    count_skipped += 1
            else:
                create_quiz(conn, title, "Imported from Excel", questions)
                count_created += 1
                print(f"Created: {title} ({len(questions)} questions)")

    print(f"Done. Created: {count_created}, Replaced: {count_replaced}, Skipped: {count_skipped}")


if __name__ == "__main__":
    main()
